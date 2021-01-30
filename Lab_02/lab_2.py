import timeit
import random
from openpyxl import Workbook

def makelist(Length):
    list1 = []
    for i in range(Length):
        list1.append(i)
    return list1
    
def copy(List):
    List.copy()

def timetest(runs, Length):
    total = 0
    list1 = []
    for _ in range(runs):
        list1 = makelist(Length)
        start = timeit.default_timer()
        copy(list1)
        end = timeit.default_timer()
        total += end - start
    return total/runs

def excel():
    filename = "Copy.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    for i in range(1000, 100000, 1000):
        total = 0
        sheet["A"+str(int(i/1000))] = i
        sheet["B"+str(int(i/1000))] = timetest(50, i)
        
    workbook.save("Copy.xlsx")

def lookups():
    L = []
    filename = "lookups_rev_2.xlsx"
    workbook = Workbook()
    sheet = workbook.active


    for i in range(1000000):
        L.append(random.randrange(500))

    x = 0

    for i in range(1000000):
        start = timeit.default_timer()
        L[i]
        end = timeit.default_timer() - start
        sheet["A"+str(i+1)] = i
        sheet["B"+str(i+1)] = end

        #print(timeit.default_timer() - start)

    workbook.save(filename=filename)


def append():
    filename = "append.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    L = []

    for i in range(1000000):
        start = timeit.default_timer()
        L.append(1)
        end = timeit.default_timer() - start
        sheet["A" + str(i+1)] = i
        sheet["B" + str(i+1)] = end

    workbook.save(filename=filename)

#append()

def multiAppend(n):
    L = []

    for i in range(n):
        L.append(i)


def multiAppendTime():
    filename = "experiments.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    cell = 1

    for i in range(1000, 100000, 1000):
        accum = 0
        for j in range(10):
            start = timeit.default_timer()
            multiAppend(i)
            end = timeit.default_timer() - start
            accum += end
        
        accum /= 10

        sheet["A" + str(cell)] = i
        sheet["B" + str(cell)] = accum

        cell += 1
        

    workbook.save(filename = filename)
    

    

#excel()
#lookups()
multiAppendTime()
